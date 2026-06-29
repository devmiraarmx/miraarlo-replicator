import requests as req_lib
from datetime import datetime, timedelta
from flask import render_template, request, jsonify, send_file
from flask_login import login_required, current_user
from app.editor import editor_bp
from app.editor.meli import MeliClient, BROWSER_HEADERS
from app.editor.claude_helper import ClaudeHelper
from app.utils.crypto import encrypt_token, decrypt_token

claude = ClaudeHelper()


class MeliClientDB(MeliClient):
    """MeliClient que lee y persiste tokens cifrados en la DB del usuario."""

    def __init__(self, user):
        super().__init__()
        self._user = user
        self.access_token = decrypt_token(user.ml_access_token) if user.ml_access_token else ''
        self.refresh_token_val = decrypt_token(user.ml_refresh_token) if user.ml_refresh_token else ''

    def _persist_tokens(self):
        from app.extensions import db
        self._user.ml_access_token = encrypt_token(self.access_token)
        self._user.ml_refresh_token = encrypt_token(self.refresh_token_val)
        expires_in = getattr(self, 'expires_in', 21600)
        self._user.ml_token_expires_at = datetime.utcnow() + timedelta(seconds=expires_in)
        db.session.commit()

    def refresh_token(self):
        result = super().refresh_token()
        if result.get('success'):
            self._persist_tokens()
        return result

    def _reload_token(self):
        """Sobreescribe el reload desde .env — en DB los tokens ya están cargados."""
        pass


def _meli_for_user():
    """Devuelve un MeliClientDB con los tokens cifrados del usuario actual."""
    if current_user.is_authenticated and current_user.ml_access_token:
        return MeliClientDB(current_user._get_current_object())
    return MeliClient()  # sin token (editor muestra banner de conexión)


@editor_bp.route('/')
@login_required
def index():
    auth_result = request.args.get('auth')
    meli = _meli_for_user()
    token_status = "configured" if meli.has_token() else "missing"
    return render_template('editor/index.html', token_status=token_status, auth_result=auth_result)


@editor_bp.route('/extract', methods=['POST'])
@login_required
def extract():
    data = request.json
    input_value = data.get('input', '').strip()
    if not input_value:
        return jsonify({'success': False, 'error': 'Ingresa un URL o ID de Mercado Libre.'})
    meli = _meli_for_user()
    result = meli.extract_item(input_value)
    return jsonify(result)


@editor_bp.route('/enhance', methods=['POST'])
@login_required
def enhance():
    data = request.json
    result = claude.enhance(
        data.get('action'),
        data.get('title', ''),
        data.get('description', ''),
        data.get('category', '')
    )
    return jsonify(result)


@editor_bp.route('/publish', methods=['POST'])
@login_required
def publish():
    from app.models import Publication
    from app.extensions import db

    # Verificar saldo antes de publicar
    row = db.session.execute(
        db.text("SELECT available_credits FROM user_credit_balance WHERE user_id = :uid"),
        {'uid': current_user.id}
    ).fetchone()
    available = int(row[0]) if row else 0
    if available < 1:
        return jsonify({
            'success': False,
            'error': 'sin_creditos',
            'message': 'No tienes créditos disponibles. Compra un paquete para seguir publicando.',
        }), 402

    meli = _meli_for_user()
    data = request.json
    result = meli.publish_item(data)

    if result.get('success'):
        pub = Publication(
            user_id=current_user.id,
            source_mlm=data.get('mlm_id', ''),
            new_mlm=result.get('new_id', ''),
            title=(data.get('title', ''))[:100],
            category_id=data.get('category_id', ''),
            price=data.get('price'),
            status='published',
            credits_used=1,
        )
        db.session.add(pub)
        db.session.commit()

    return jsonify(result)


@editor_bp.route('/category-attributes', methods=['POST'])
@login_required
def category_attributes():
    data = request.json
    category_id = data.get('category_id', '').strip()
    existing_attrs = data.get('attributes', {})

    if not category_id:
        return jsonify({'success': False, 'error': 'Sin categoría'})

    meli = _meli_for_user()
    cat_attrs = meli.get_category_attributes(category_id)

    prefilled = {}
    for ca in cat_attrs:
        attr_id = ca['id']
        if attr_id in existing_attrs:
            ea = existing_attrs[attr_id]
            if isinstance(ea, dict):
                prefilled[attr_id] = ea.get('value_name', '')
            else:
                prefilled[attr_id] = str(ea)

    return jsonify({
        'success': True,
        'category_id': category_id,
        'attributes': cat_attrs,
        'prefilled': prefilled
    })


@editor_bp.route('/refresh-token', methods=['POST'])
@login_required
def refresh_token():
    meli = _meli_for_user()
    return jsonify(meli.refresh_token())


@editor_bp.route('/me')
@login_required
def me():
    meli = _meli_for_user()
    user = meli.get_me()
    if user:
        return jsonify({'success': True, 'id': user.get('id'), 'nickname': user.get('nickname')})
    return jsonify({'success': False, 'error': 'Token invalido'})


@editor_bp.route('/predict-category', methods=['POST'])
@login_required
def predict_category():
    title = request.json.get('title', '')
    if not title:
        return jsonify({'success': False, 'error': 'Sin titulo'})
    meli = _meli_for_user()
    cat = meli.predict_category(title)
    if cat:
        return jsonify({'success': True, 'category_id': cat})
    return jsonify({'success': False, 'error': 'No se pudo predecir la categoria'})


@editor_bp.route('/category-path', methods=['POST'])
@login_required
def category_path():
    category_id = request.json.get('category_id', '').strip()
    if not category_id:
        return jsonify({'success': False, 'error': 'Sin categoría'})
    try:
        r = req_lib.get(
            f'https://api.mercadolibre.com/categories/{category_id}',
            headers={'Accept': 'application/json'}, timeout=8
        )
        if r.status_code == 200:
            data = r.json()
            path = [p.get('name', '') for p in data.get('path_from_root', [])]
            return jsonify({
                'success': True,
                'category_id': category_id,
                'name': data.get('name', ''),
                'path': path,
                'breadcrumb': ' > '.join(path)
            })
        return jsonify({'success': False, 'error': f'Error {r.status_code}'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@editor_bp.route('/export-excel', methods=['POST'])
@login_required
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    data = request.json

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listing Miraarlo"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A1D27")
    accent_fill = PatternFill("solid", fgColor="FFE600")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.merge_cells('A1:C1')
    ws['A1'] = f"Miraarlo Replicator — {data.get('mlm_id', '')} — {data.get('title', '')[:40]}"
    ws['A1'].font = Font(bold=True, size=13, color="111111")
    ws['A1'].fill = accent_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ['Campo', 'Valor', 'Notas']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    fields = [
        ('ID Original',       data.get('mlm_id', ''),              'ID de la publicación fuente'),
        ('Título',            data.get('title', ''),               'Máx. 60 caracteres'),
        ('Precio (MXN)',      data.get('price', ''),               'Sin centavos si es redondo'),
        ('Moneda',            data.get('currency', 'MXN'),         ''),
        ('Categoría ID',      data.get('category_id', ''),         'No modificar'),
        ('Condición',         data.get('condition', 'new'),         'new / used'),
        ('Stock',             data.get('available_quantity', 1),   ''),
        ('Tipo de anuncio',   data.get('listing_type', ''),        ''),
        ('Descripción',       data.get('description', ''),         'Texto plano'),
    ]

    for k, v in (data.get('dynamic_attrs') or {}).items():
        if v:
            fields.append((f'Atributo: {k}', v, ''))

    for k, v in (data.get('attributes') or {}).items():
        fields.append((f'Atributo: {k}', v if isinstance(v, str) else (v.get('value_name', '') if isinstance(v, dict) else str(v)), ''))

    for i, url in enumerate(data.get('photos', []), 1):
        fields.append((f'Foto {i}', url, 'URL imagen'))

    row = 3
    for campo, valor, nota in fields:
        ws.cell(row=row, column=1, value=campo).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F5F5F5")
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=1).border = border

        val_cell = ws.cell(row=row, column=2, value=str(valor) if valor else '')
        val_cell.alignment = left
        val_cell.border = border
        if campo == 'Descripción':
            ws.row_dimensions[row].height = max(60, min(len(str(valor or '')) / 2, 200))

        ws.cell(row=row, column=3, value=nota).font = Font(italic=True, color="888888", size=9)
        ws.cell(row=row, column=3).alignment = left
        ws.cell(row=row, column=3).border = border
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 25

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"miraarlo_{data.get('mlm_id', 'listing')}.xlsx"
    return send_file(buf,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@editor_bp.route('/download-photos', methods=['POST'])
@login_required
def download_photos():
    import zipfile
    from io import BytesIO

    data = request.json
    photos = data.get('photos', [])
    mlm_id = data.get('mlm_id', 'listing')

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, url in enumerate(photos, 1):
            try:
                r = req_lib.get(url, headers=BROWSER_HEADERS, timeout=15)
                if r.status_code == 200:
                    ext = 'jpg' if 'jpg' in url.lower() else 'png'
                    zf.writestr(f"{mlm_id}_foto_{i:02d}.{ext}", r.content)
            except Exception:
                continue

    buf.seek(0)
    return send_file(buf, mimetype='application/zip',
                     as_attachment=True, download_name=f"fotos_{mlm_id}.zip")


@editor_bp.route('/fill-template', methods=['POST'])
@login_required
def fill_template():
    import openpyxl
    import json as _json
    from io import BytesIO

    if 'template' not in request.files:
        return jsonify({'success': False, 'error': 'No se recibió archivo de plantilla'}), 400

    template_file = request.files['template']
    item_json = request.form.get('item_data', '{}')

    try:
        item = _json.loads(item_json)
    except Exception:
        return jsonify({'success': False, 'error': 'JSON de item inválido'}), 400

    try:
        wb = openpyxl.load_workbook(template_file)
    except Exception as e:
        return jsonify({'success': False, 'error': f'No se pudo leer la plantilla: {e}'}), 400

    skip = {'ayuda', 'extra info'}
    product_sheet = next((wb[s] for s in wb.sheetnames if s.lower() not in skip), None)
    if not product_sheet:
        return jsonify({'success': False, 'error': 'No se encontró hoja de producto en la plantilla'}), 400

    ws = product_sheet
    HEADER_ROW = 3
    FIRST_DATA = 8

    col_map = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=HEADER_ROW, column=col).value
        if raw:
            col_map[str(raw).lower().strip()] = col

    def find_col(*keywords):
        for key in keywords:
            for name, idx in col_map.items():
                if key.lower() in name:
                    return idx
        return None

    condition_raw = item.get('condition', 'new')
    condition_ml = 'Nuevo' if condition_raw == 'new' else 'Usado'
    listing_type_raw = item.get('listing_type', 'gold_special')
    listing_ml = 'Premium' if listing_type_raw == 'gold_pro' else 'Clásica'
    photos = item.get('photos', [])
    photos_str = ', '.join(p for p in photos if p and str(p).startswith('http'))
    dynamic_attrs = item.get('dynamic_attrs') or {}

    fill_map = {
        find_col('título', 'titulo', 'title'):            item.get('title', '')[:60],
        find_col('condición', 'condicion', 'condition'):  condition_ml,
        find_col('fotos', 'photos', 'imagen'):            photos_str,
        find_col('stock'):                                item.get('available_quantity', 1),
        find_col('precio', 'price'):                      item.get('price', 0),
        find_col('moneda', 'currency'):                   '$',
        find_col('descripción', 'descripcion'):           item.get('description', ''),
        find_col('tipo de publicación', 'tipo publicac'): listing_ml,
        find_col('forma de envío', 'forma de envio'):     'Mercado Envíos',
        find_col('costo de envío', 'costo de envio'):     'A cargo del comprador',
        find_col('tipo de garantía', 'tipo de garantia'): 'Garantía del vendedor',
        find_col('tiempo de garantía', 'tiempo de garant'): 3,
        find_col('unidad de tiempo'):                     'meses',
        find_col('marca'):                                dynamic_attrs.get('BRAND', ''),
        find_col('modelo'):                               dynamic_attrs.get('MODEL', ''),
        find_col('material'):                             dynamic_attrs.get('MATERIAL', ''),
        find_col('color'):                                dynamic_attrs.get('COLOR', ''),
    }

    fill_map = {k: v for k, v in fill_map.items() if k and v not in ('', None, 0)}

    for col_idx, value in fill_map.items():
        try:
            ws.cell(row=FIRST_DATA, column=col_idx, value=value)
        except AttributeError:
            pass

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    mlm_id = item.get('mlm_id', 'item')
    filename = f"ML_plantilla_{mlm_id}_lista.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
