from flask import Flask, render_template, request, jsonify, redirect
import os, requests as req_lib
from dotenv import load_dotenv
from werkzeug.middleware.proxy_fix import ProxyFix
from meli import MeliClient, BROWSER_HEADERS
from claude_helper import ClaudeHelper

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'miraarlo-secret-2024')
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1, x_prefix=1)

meli = MeliClient()
claude = ClaudeHelper()


def get_callback_url():
    proto = request.headers.get('X-Forwarded-Proto', 'http')
    host  = request.headers.get('X-Forwarded-Host', request.host)
    return f"{proto}://{host}/auth/callback"


@app.route('/')
def index():
    auth_result = request.args.get('auth')
    token_status = "configured" if meli.has_token() else "missing"
    return render_template('index.html', token_status=token_status, auth_result=auth_result)


@app.route('/extract', methods=['POST'])
def extract():
    data = request.json
    input_value = data.get('input', '').strip()
    if not input_value:
        return jsonify({'success': False, 'error': 'Ingresa un URL o ID de Mercado Libre.'})
    result = meli.extract_item(input_value)
    return jsonify(result)


@app.route('/enhance', methods=['POST'])
def enhance():
    data = request.json
    result = claude.enhance(data.get('action'), data.get('title',''), data.get('description',''), data.get('category',''))
    return jsonify(result)


@app.route('/publish', methods=['POST'])
def publish():
    result = meli.publish_item(request.json)
    return jsonify(result)


@app.route('/category-attributes', methods=['POST'])
def category_attributes():
    """Devuelve los atributos obligatorios de una categoría + pre-llena con datos del item."""
    data = request.json
    category_id = data.get('category_id', '').strip()
    existing_attrs = data.get('attributes', {})

    if not category_id:
        return jsonify({'success': False, 'error': 'Sin categoría'})

    cat_attrs = meli.get_category_attributes(category_id)

    # Pre-llenar con atributos existentes del item
    prefilled = {}
    for ca in cat_attrs:
        attr_id = ca['id']
        if attr_id in existing_attrs:
            ea = existing_attrs[attr_id]
            if isinstance(ea, dict):
                prefilled[attr_id] = ea.get('value_name', '')
            else:
                prefilled[attr_id] = str(ea)

    return jsonify({
        'success': True,
        'category_id': category_id,
        'attributes': cat_attrs,
        'prefilled': prefilled
    })


@app.route('/auth')
def auth():
    base_url = request.args.get('base_url', '').rstrip('/')
    if not base_url:
        base_url = request.host_url.rstrip('/')
    callback_url = base_url + '/auth/callback'
    from flask import session
    session['callback_url'] = callback_url
    print(f"[AUTH] Usando callback_url: {callback_url}")
    return redirect(meli.get_auth_url(callback_url))


@app.route('/auth/callback')
def auth_callback():
    from flask import session
    code = request.args.get('code')
    if not code:
        return redirect('/?auth=error')
    callback_url = session.get('callback_url', request.host_url.rstrip('/') + '/auth/callback')
    print(f"[AUTH CALLBACK] code recibido, usando callback_url: {callback_url}")
    result = meli.exchange_code(code, callback_url)
    print(f"[AUTH CALLBACK] resultado: {result}")
    return redirect('/?auth=success' if result.get('success') else '/?auth=error')


@app.route('/refresh-token', methods=['POST'])
def refresh_token():
    return jsonify(meli.refresh_token())


@app.route('/export-excel', methods=['POST'])
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from flask import send_file

    data = request.json

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Listing Miraarlo"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A1D27")
    accent_fill = PatternFill("solid", fgColor="FFE600")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    ws.merge_cells('A1:C1')
    ws['A1'] = f"Miraarlo Replicator — {data.get('mlm_id', '')} — {data.get('title', '')[:40]}"
    ws['A1'].font = Font(bold=True, size=13, color="111111")
    ws['A1'].fill = accent_fill
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 28

    headers = ['Campo', 'Valor', 'Notas']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    fields = [
        ('ID Original',       data.get('mlm_id', ''),              'ID de la publicación fuente'),
        ('Título',            data.get('title', ''),               'Máx. 60 caracteres'),
        ('Precio (MXN)',      data.get('price', ''),               'Sin centavos si es redondo'),
        ('Moneda',            data.get('currency', 'MXN'),         ''),
        ('Categoría ID',      data.get('category_id', ''),         'No modificar'),
        ('Condición',         data.get('condition', 'new'),         'new / used'),
        ('Stock',             data.get('available_quantity', 1),   ''),
        ('Tipo de anuncio',   data.get('listing_type', ''),        ''),
        ('Descripción',       data.get('description', ''),         'Texto plano'),
    ]

    # Atributos dinámicos
    for k, v in (data.get('dynamic_attrs') or {}).items():
        if v:
            fields.append((f'Atributo: {k}', v, ''))

    # Atributos legacy
    for k, v in (data.get('attributes') or {}).items():
        fields.append((f'Atributo: {k}', v if isinstance(v, str) else (v.get('value_name','') if isinstance(v,dict) else str(v)), ''))

    for i, url in enumerate(data.get('photos', []), 1):
        fields.append((f'Foto {i}', url, 'URL imagen'))

    row = 3
    for campo, valor, nota in fields:
        ws.cell(row=row, column=1, value=campo).font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F5F5F5")
        ws.cell(row=row, column=1).alignment = left
        ws.cell(row=row, column=1).border = border

        val_cell = ws.cell(row=row, column=2, value=str(valor) if valor else '')
        val_cell.alignment = left
        val_cell.border = border
        if campo == 'Descripción':
            ws.row_dimensions[row].height = max(60, min(len(str(valor or '')) / 2, 200))

        ws.cell(row=row, column=3, value=nota).font = Font(italic=True, color="888888", size=9)
        ws.cell(row=row, column=3).alignment = left
        ws.cell(row=row, column=3).border = border
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 70
    ws.column_dimensions['C'].width = 25

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"miraarlo_{data.get('mlm_id', 'listing')}.xlsx"
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/download-photos', methods=['POST'])
def download_photos():
    import zipfile
    from io import BytesIO
    from flask import send_file

    data = request.json
    photos = data.get('photos', [])
    mlm_id = data.get('mlm_id', 'listing')

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for i, url in enumerate(photos, 1):
            try:
                r = req_lib.get(url, headers=BROWSER_HEADERS, timeout=15)
                if r.status_code == 200:
                    ext = 'jpg' if 'jpg' in url.lower() else 'png'
                    zf.writestr(f"{mlm_id}_foto_{i:02d}.{ext}", r.content)
            except Exception:
                continue

    buf.seek(0)
    return send_file(buf, mimetype='application/zip',
                     as_attachment=True, download_name=f"fotos_{mlm_id}.zip")


@app.route('/me')
def me():
    user = meli.get_me()
    if user:
        return jsonify({'success': True, 'id': user.get('id'), 'nickname': user.get('nickname')})
    return jsonify({'success': False, 'error': 'Token invalido'})


@app.route('/predict-category', methods=['POST'])
def predict_category():
    title = request.json.get('title', '')
    if not title:
        return jsonify({'success': False, 'error': 'Sin titulo'})
    cat = meli.predict_category(title)
    if cat:
        return jsonify({'success': True, 'category_id': cat})
    return jsonify({'success': False, 'error': 'No se pudo predecir la categoria'})


@app.route('/category-path', methods=['POST'])
def category_path():
    """Devuelve el breadcrumb de una categoría de ML."""
    category_id = request.json.get('category_id', '').strip()
    if not category_id:
        return jsonify({'success': False, 'error': 'Sin categoría'})
    try:
        r = req_lib.get(
            f'https://api.mercadolibre.com/categories/{category_id}',
            headers={'Accept': 'application/json'}, timeout=8
        )
        if r.status_code == 200:
            data = r.json()
            path = [p.get('name', '') for p in data.get('path_from_root', [])]
            return jsonify({
                'success': True,
                'category_id': category_id,
                'name': data.get('name', ''),
                'path': path,
                'breadcrumb': ' > '.join(path)
            })
        return jsonify({'success': False, 'error': f'Error {r.status_code}'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/fill-template', methods=['POST'])
def fill_template():
    import openpyxl
    from io import BytesIO
    from flask import send_file

    if 'template' not in request.files:
        return jsonify({'success': False, 'error': 'No se recibió archivo de plantilla'}), 400

    template_file = request.files['template']
    item_json     = request.form.get('item_data', '{}')

    try:
        import json as _json
        item = _json.loads(item_json)
    except Exception:
        return jsonify({'success': False, 'error': 'JSON de item inválido'}), 400

    try:
        wb = openpyxl.load_workbook(template_file)
    except Exception as e:
        return jsonify({'success': False, 'error': f'No se pudo leer la plantilla: {e}'}), 400

    skip = {'ayuda', 'extra info'}
    product_sheet = next(
        (wb[s] for s in wb.sheetnames if s.lower() not in skip), None
    )
    if not product_sheet:
        return jsonify({'success': False, 'error': 'No se encontró hoja de producto en la plantilla'}), 400

    ws = product_sheet
    HEADER_ROW  = 3
    FIRST_DATA  = 8

    col_map = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=HEADER_ROW, column=col).value
        if raw:
            col_map[str(raw).lower().strip()] = col

    def find_col(*keywords):
        for key in keywords:
            for name, idx in col_map.items():
                if key.lower() in name:
                    return idx
        return None

    condition_raw = item.get('condition', 'new')
    condition_ml  = 'Nuevo' if condition_raw == 'new' else 'Usado'

    listing_type_raw = item.get('listing_type', 'gold_special')
    listing_ml = 'Premium' if listing_type_raw == 'gold_pro' else 'Clásica'

    photos = item.get('photos', [])
    photos_str = ', '.join(p for p in photos if p and str(p).startswith('http'))

    # Obtener valores de atributos dinámicos
    dynamic_attrs = item.get('dynamic_attrs') or {}

    fill_map = {
        find_col('título', 'titulo', 'title'):           item.get('title', '')[:60],
        find_col('condición', 'condicion', 'condition'): condition_ml,
        find_col('fotos', 'photos', 'imagen'):           photos_str,
        find_col('stock'):                               item.get('available_quantity', 1),
        find_col('precio', 'price'):                     item.get('price', 0),
        find_col('moneda', 'currency'):                  '$',
        find_col('descripción', 'descripcion'):          item.get('description', ''),
        find_col('tipo de publicación', 'tipo publicac'): listing_ml,
        find_col('forma de envío', 'forma de envio'):    'Mercado Envíos',
        find_col('costo de envío', 'costo de envio'):    'A cargo del comprador',
        find_col('tipo de garantía', 'tipo de garantia'): 'Garantía del vendedor',
        find_col('tiempo de garantía', 'tiempo de garant'): 3,
        find_col('unidad de tiempo'):                    'meses',
        find_col('marca'):                               dynamic_attrs.get('BRAND', ''),
        find_col('modelo'):                              dynamic_attrs.get('MODEL', ''),
        find_col('material'):                            dynamic_attrs.get('MATERIAL', ''),
        find_col('color'):                               dynamic_attrs.get('COLOR', ''),
    }

    fill_map = {k: v for k, v in fill_map.items() if k and v not in ('', None, 0)}

    for col_idx, value in fill_map.items():
        try:
            ws.cell(row=FIRST_DATA, column=col_idx, value=value)
        except AttributeError:
            pass

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    mlm_id   = item.get('mlm_id', 'item')
    filename = f"ML_plantilla_{mlm_id}_lista.xlsx"
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


if __name__ == '__main__':
    print("\n🟡 Miraarlo Replicator corriendo en http://localhost:5000\n")
    app.run(debug=True, port=5000, use_reloader=False)
